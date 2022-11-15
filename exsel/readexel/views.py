from django import forms
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.contrib.auth.models import User

# from .models import User


import openpyxl
from .models import Book
import csv

# class NewLotForm(forms.Form):
#     file = forms.FileField(widget=forms.ClearableFileInput(attrs={'multiple': True}))

class NewLotForm(forms.Form):
    some_text = forms.CharField(label="data:")


# Create your views here.
# def index(request):
#     if request.method == "POST":
#         file = request.FILES["file"]  # form = NewLotForm(request.POST, request.FILES)
#         # file = NewLotForm(request.POST, request.FILES)
#         wookbook = openpyxl.load_workbook(file)
#         # Define variable to read the active sheet:
#         worksheet = wookbook.active
#         print(worksheet.max_row)
#         print(worksheet.max_column)
#         # Iterate the loop to read the cell values
#         for rows in worksheet.iter_rows(2, worksheet.max_row):
#             print(rows[0].value)
#             try:
#                 if rows[0].value !=None and rows[1].value !=None and rows[2].value !=None:
#                     book = Book.objects.create(photo_link=rows[0].value, name_of_event=rows[1].value, date_of_event=rows[2].value,
#                                                description=rows[3].value, start_time=rows[4].value, end_time=rows[5].value,
#                                                location_name=rows[6].value, first_event_category=rows[7].value,
#                                                second_category=rows[8].value, free_or_Paid=rows[9].value,
#                                                ticket_site_link=rows[10].value,
#                                                )
#                     book.save()
#                 else:
#                     break
#
#             except:
#                 return HttpResponse("file not correct")
#
#
#         return HttpResponse("your model is already in the database go to https://yan3608.pythonanywhere.com/admin (login: yan password: 390871) and make sure. And remove empty values rom the table. there are a thousand empty values")
#     else:
#         return render(request, "readexel/index.html", {'form': NewLotForm})

def index(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            some_text = request.POST["some_text"]
            return render(request, "readexel/response.html", {'some_text': some_text})
        else:
            return render(request, "readexel/Workspace.html", {'form': NewLotForm})

    else:
        return render(request, "readexel/index.html")



def login_view(request):
    if request.method == "POST":

        # Attempt to sign user in
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        # Check if authentication successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("index"))
        else:
            return render(request, "readexel/login.html", {
                "message": "Invalid username and/or password."
            })
    else:
        return render(request, "readexel/login.html")


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("index"))


def register(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "readexel/register.html", {
                "message": "Passwords must match."
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()
        except IntegrityError:
            return render(request, "readexel/register.html", {
                "message": "Username already taken."
            })
        login(request, user)
        return HttpResponseRedirect(reverse("index"))
    else:
        return render(request, "readexel/register.html")
